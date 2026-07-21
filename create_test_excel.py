import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ============ SHEET 1: LOGIN PAGE ============
ws1 = wb.active
ws1.title = "Login Page"

headers = ["Test Case ID", "Module", "Test Scenario", "Test Case Description", "Test Type",
           "Preconditions", "Test Steps", "Expected Result", "Actual Result", "Status (Pass/Fail)", "Remarks"]

header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
body_align = Alignment(vertical="top", wrap_text=True)

for col, h in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

login_cases = [
    # (id, scenario, description, test_type, preconditions, steps, expected)
    ("LGN_001", "UI Verification", "Verify login page UI elements", "UI/UX Testing",
     "Application URL is accessible",
     "1. Navigate to the login page\n2. Observe the page layout",
     "All UI elements (username, password fields, login button, forgot password link) are properly aligned and displayed"),

    ("LGN_002", "Placeholder & Validation Text", "Verify placeholder texts and validation messages", "UI/UX Testing",
     "Login page is open",
     "1. Check placeholder text in username field\n2. Check placeholder text in password field",
     "Placeholder text is displayed correctly as per design spec"),

    ("LGN_003", "Empty Fields Validation", "Verify login with empty fields", "Negative Testing",
     "Login page is open",
     "1. Leave username and password empty\n2. Click the Login button",
     "Proper validation message should be displayed: 'Please enter username' and 'Please enter password'"),

    ("LGN_004", "Valid Credentials Login", "Verify login with valid credentials", "Functional Testing",
     "Valid username and password available",
     "1. Enter valid username\n2. Enter valid password\n3. Click Login button",
     "User should be redirected to the dashboard/homepage successfully"),

    ("LGN_005", "Invalid Username", "Verify login with invalid username and valid password", "Negative Testing",
     "Login page is open",
     "1. Enter invalid username\n2. Enter valid password\n3. Click Login button",
     "Error message: 'Invalid username or password' should be displayed"),

    ("LGN_006", "Invalid Password", "Verify login with valid username and invalid password", "Negative Testing",
     "Login page is open",
     "1. Enter valid username\n2. Enter invalid password\n3. Click Login button",
     "Error message: 'Invalid username or password' should be displayed"),

    ("LGN_007", "Both Invalid", "Verify login with both invalid credentials", "Negative Testing",
     "Login page is open",
     "1. Enter invalid username\n2. Enter invalid password\n3. Click Login button",
     "Error message: 'Invalid username or password' should be displayed"),

    ("LGN_008", "SQL Injection", "Verify login with SQL injection in username field", "Security Testing",
     "Login page is open",
     "1. Enter ' OR 1=1 -- in username\n2. Enter any password\n3. Click Login button",
     "SQL injection should be prevented; error message should be shown"),

    ("LGN_009", "Password Masking", "Verify password is masked/hidden", "Security Testing",
     "Login page is open",
     "1. Enter text in password field\n2. Observe the field display",
     "Password characters should be displayed as dots/asterisks"),

    ("LGN_010", "Password Visibility Toggle", "Verify password show/hide toggle", "UI/UX Testing",
     "Login page is open",
     "1. Enter password\n2. Click the eye/toggle icon\n3. Verify password visibility",
     "Password should toggle between masked and visible states"),

    ("LGN_011", "Remember Me", "Verify Remember Me checkbox functionality", "Functional Testing",
     "Login page is open",
     "1. Check 'Remember Me' checkbox\n2. Enter valid credentials\n3. Login and logout\n4. Verify username field is pre-filled",
     "Username should be remembered/populated on the next login attempt"),

    ("LGN_012", "Forgot Password Link", "Verify Forgot Password link redirects correctly", "Functional Testing",
     "Login page is open",
     "1. Click 'Forgot Password' link",
     "User should be redirected to the Forgot Password / Reset Password page"),

    ("LGN_013", "Case Sensitivity - Username", "Verify username field is case-sensitive or not", "Functional Testing",
     "Valid credentials available",
     "1. Enter username in UPPERCASE\n2. Enter valid password\n3. Click Login",
     "Application behavior based on requirement (should note if case-sensitive)"),

    ("LGN_014", "Copy/Paste in Password", "Verify copy-paste in password field", "Functional Testing",
     "Login page is open",
     "1. Right-click on password field\n2. Try to copy text",
     "Copy functionality should be disabled for the password field"),

    ("LGN_015", "Keyboard Navigation", "Verify Tab and Enter key navigation", "Usability Testing",
     "Login page is open",
     "1. Press Tab key from username\n2. Press Tab again\n3. Press Enter on Login button",
     "Focus should move sequentially: username -> password -> login button. Enter should submit the form"),

    ("LGN_016", "Session Timeout", "Verify session timeout redirects to login", "Security Testing",
     "User is logged in",
     "1. Stay idle until session expires\n2. Try to access any page",
     "User should be redirected to login page with session timeout message"),

    ("LGN_017", "Multiple Failed Attempts", "Verify account lockout after multiple failed attempts", "Security Testing",
     "Login page is open",
     "1. Enter invalid credentials 3-5 times\n2. Try with valid credentials",
     "Account should be locked after maximum failed attempts; user should see lockout message"),

    ("LGN_018", "Logged-In User Redirect", "Verify already logged-in user is redirected away from login page", "Functional Testing",
     "User is already logged in",
     "1. Manually navigate to login page URL\n2. Observe behavior",
     "User should be auto-redirected to dashboard/homepage"),

    ("LGN_019", "Loading State", "Verify loading indicator during login", "UI/UX Testing",
     "Login page is open",
     "1. Enter valid credentials\n2. Click Login\n3. Observe loading state",
     "Loading spinner/progress indicator should appear while login request is in progress"),

    ("LGN_020", "API Response Error Handling", "Verify error handling for server errors (5xx)", "Integration Testing",
     "Server is down / Network is slow",
     "1. Enter valid credentials\n2. Click Login\n3. Observe error handling",
     "User-friendly error message should be displayed: 'Server error. Please try again later'"),

    ("LGN_021", "Special Characters in Fields", "Verify special characters handling in username/password", "Security Testing",
     "Login page is open",
     "1. Enter HTML/script tags in username (<script>alert('xss')</script>)\n2. Enter password\n3. Click Login",
     "XSS should be prevented; input should be sanitized"),

    ("LGN_022", "Browser Back/Forward", "Verify browser navigation after login", "Functional Testing",
     "User has logged in",
     "1. Click browser Back button\n2. Click browser Forward button",
     "Back button should not go to login page; forward should work normally"),

    ("LGN_023", "Response Time", "Verify login page loads within acceptable time", "Performance Testing",
     "Application is running",
     "1. Navigate to login page\n2. Measure page load time",
     "Login page should load within 3 seconds"),

    ("LGN_024", "Mobile Responsiveness", "Verify login page on mobile viewport", "Responsive Testing",
     "Application is accessible on mobile browser",
     "1. Open login page on mobile device\n2. Observe layout and usability",
     "Login page should be responsive; fields and buttons should be properly sized for touch"),
]

for i, (tid, scenario, desc, ttype, precond, steps, expected) in enumerate(login_cases, 2):
    row = [tid, "Login Page", scenario, desc, ttype, precond, steps, expected, "", "", ""]
    for col, val in enumerate(row, 1):
        cell = ws1.cell(row=i, column=col, value=val)
        cell.alignment = body_align
        cell.border = thin_border

# Column widths
col_widths = [12, 14, 28, 40, 18, 28, 40, 40, 18, 18, 20]
for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w


# ============ SHEET 2: WORKSPACE ============
ws2 = wb.create_sheet("Workspace")

for col, h in enumerate(headers, 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    cell.alignment = header_align
    cell.border = thin_border

workspace_cases = [
    ("WRK_001", "UI Verification", "Verify workspace listing page UI", "UI/UX Testing",
     "User is logged in",
     "1. Navigate to Workspace module\n2. Observe the workspace list view",
     "All workspace cards/list items display properly with name, status, and action buttons"),

    ("WRK_002", "Create Workspace", "Verify creating a new workspace", "Functional Testing",
     "User has admin privileges",
     "1. Click 'Create Workspace'\n2. Enter workspace name, description\n3. Select workspace type\n4. Click Save",
     "New workspace should be created successfully and shown in the list"),

    ("WRK_003", "Create - Empty Name", "Verify validation when workspace name is empty", "Negative Testing",
     "User is on Create Workspace form",
     "1. Leave Name field empty\n2. Enter description\n3. Click Save",
     "Validation error: 'Workspace name is required' should be displayed"),

    ("WRK_004", "Create - Duplicate Name", "Verify creating workspace with an existing name", "Negative Testing",
     "A workspace with same name already exists",
     "1. Enter existing workspace name\n2. Enter description\n3. Click Save",
     "Error message: 'Workspace name already exists. Please use a different name'"),

    ("WRK_005", "Edit Workspace", "Verify editing an existing workspace", "Functional Testing",
     "At least one workspace exists",
     "1. Click Edit on a workspace\n2. Modify name/description\n3. Click Update",
     "Workspace details should be updated successfully"),

    ("WRK_006", "Delete Workspace", "Verify deleting a workspace", "Functional Testing",
     "At least one workspace exists",
     "1. Click Delete on a workspace\n2. Confirm deletion",
     "Workspace should be soft-deleted / moved to trash; removed from active list"),

    ("WRK_007", "Delete - Cancel", "Verify cancellation of workspace deletion", "Functional Testing",
     "At least one workspace exists",
     "1. Click Delete\n2. Click Cancel on confirmation dialog",
     "Workspace should NOT be deleted; it should remain in the list"),

    ("WRK_008", "Search Workspace", "Verify workspace search functionality", "Functional Testing",
     "Multiple workspaces exist",
     "1. Enter a keyword in search bar\n2. Observe filtered results",
     "Workspaces matching the keyword should be displayed in real-time"),

    ("WRK_009", "Filter by Status", "Verify filtering workspaces by status (Active/Inactive/Archived)", "Functional Testing",
     "Workspaces with different statuses exist",
     "1. Click on filter dropdown\n2. Select 'Active'\n3. Observe results",
     "Only active workspaces should be displayed"),

    ("WRK_010", "Pagination", "Verify pagination when workspaces exceed page limit", "Functional Testing",
     "More than N (e.g., 20) workspaces exist",
     "1. Navigate to workspace list\n2. Scroll to bottom\n3. Click page numbers",
     "Pagination should work; each page shows N records"),

    ("WRK_011", "Change Workspace Owner", "Verify transferring workspace ownership", "Functional Testing",
     "Workspace exists with owner assigned",
     "1. Open workspace settings\n2. Select a different owner\n3. Save changes",
     "Ownership should be transferred successfully; new owner gets control"),

    ("WRK_012", "Archive Workspace", "Verify archiving a workspace", "Functional Testing",
     "At least one active workspace exists",
     "1. Click 'Archive' on a workspace\n2. Confirm archiving",
     "Workspace should be archived; moved to Archived tab/filter"),

    ("WRK_013", "Restore Workspace", "Verify restoring an archived workspace", "Functional Testing",
     "At least one archived workspace exists",
     "1. Go to Archived tab\n2. Click 'Restore' on an archived workspace",
     "Workspace should be restored to active status"),

    ("WRK_014", "Workspace Settings - General", "Verify workspace general settings page", "Functional Testing",
     "User is workspace admin",
     "1. Navigate to Workspace Settings\n2. Modify general settings (timezone, language)\n3. Save",
     "General settings should be saved and applied for the workspace"),

    ("WRK_015", "Permission/Role Assignment", "Verify assigning roles within workspace", "Functional Testing",
     "Workspace exists with multiple members",
     "1. Go to Workspace Members\n2. Select a member\n3. Change role (Admin/Member/Viewer)\n4. Save",
     "Role should be updated; member's permissions should reflect the new role"),

    ("WRK_016", "Bulk Actions", "Verify bulk select and action on workspaces", "Functional Testing",
     "Multiple workspaces exist",
     "1. Select multiple workspaces via checkboxes\n2. Choose bulk action (Delete/Archive)\n3. Confirm",
     "Bulk action should apply to all selected workspaces"),

    ("WRK_017", "Audit Log", "Verify workspace audit trail/logs", "Security Testing",
     "Workspace has user activity",
     "1. Navigate to Workspace Audit Log\n2. Observe activity entries",
     "All create/edit/delete actions should be logged with user, timestamp, and details"),

    ("WRK_018", "Export Workspace Data", "Verify exporting workspace data", "Functional Testing",
     "Workspace contains data",
     "1. Go to Workspace Export\n2. Select format (CSV/Excel)\n3. Click Export",
     "Workspace data should be exported and downloaded in the selected format"),

    ("WRK_019", "Access Control - Non-Member", "Verify non-member cannot access workspace", "Security Testing",
     "User is NOT a member of the workspace",
     "1. Try to access workspace directly via URL",
     "403 Forbidden or workspace not visible; user should not have access"),

    ("WRK_020", "Performance - Large List", "Verify workspace list loads with many workspaces", "Performance Testing",
     "100+ workspaces exist",
     "1. Navigate to workspace list\n2. Measure loading time",
     "Workspace list should load within 5 seconds even with 100+ items"),
]

for i, (tid, scenario, desc, ttype, precond, steps, expected) in enumerate(workspace_cases, 2):
    row = [tid, "Workspace", scenario, desc, ttype, precond, steps, expected, "", "", ""]
    for col, val in enumerate(row, 1):
        cell = ws2.cell(row=i, column=col, value=val)
        cell.alignment = body_align
        cell.border = thin_border

for i, w in enumerate(col_widths, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w


# ============ SHEET 3: STAFFS ============
ws3 = wb.create_sheet("Staffs")

for col, h in enumerate(headers, 1):
    cell = ws3.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")
    cell.alignment = header_align
    cell.border = thin_border

staff_cases = [
    ("STF_001", "UI Verification", "Verify staff listing page UI", "UI/UX Testing",
     "User is logged in",
     "1. Navigate to Staff module\n2. Observe staff list view",
     "Staff list displays correctly with name, email, role, status, and actions"),

    ("STF_002", "Add Staff", "Verify adding a new staff member", "Functional Testing",
     "User has admin privileges",
     "1. Click 'Add Staff'\n2. Fill in name, email, phone, role\n3. Click Save",
     "New staff member should be added and appear in the staff list"),

    ("STF_003", "Add Staff - Missing Required Fields", "Verify validation when required fields are empty", "Negative Testing",
     "User is on Add Staff form",
     "1. Leave Name and Email empty\n2. Click Save",
     "Validation error messages should appear for required fields"),

    ("STF_004", "Add Staff - Duplicate Email", "Verify adding staff with existing email", "Negative Testing",
     "Email is already registered",
     "1. Enter email that already exists\n2. Fill other fields\n3. Click Save",
     "Error: 'Email already exists' should be displayed"),

    ("STF_005", "Add Staff - Invalid Email Format", "Verify email format validation", "Negative Testing",
     "User is on Add Staff form",
     "1. Enter invalid email format ('test@' or 'test.com')\n2. Click Save",
     "Validation: 'Please enter a valid email address'"),

    ("STF_006", "Edit Staff Details", "Verify editing staff information", "Functional Testing",
     "At least one staff member exists",
     "1. Click Edit on a staff member\n2. Modify name/email/phone\n3. Click Update",
     "Staff details should be updated successfully"),

    ("STF_007", "Delete Staff", "Verify deleting a staff member", "Functional Testing",
     "At least one staff member exists",
     "1. Click Delete on a staff member\n2. Confirm deletion",
     "Staff member should be deactivated/removed from active list"),

    ("STF_008", "Search Staff", "Verify staff search functionality", "Functional Testing",
     "Multiple staff members exist",
     "1. Enter name/email in search bar\n2. Observe results",
     "Staff matching the search keyword should appear"),

    ("STF_009", "Filter by Role", "Verify filtering staff by role (Admin/Manager/Staff)", "Functional Testing",
     "Staff with different roles exist",
     "1. Select a role filter\n2. Observe results",
     "Only staff with the selected role should be displayed"),

    ("STF_010", "Filter by Status", "Verify filtering staff by status (Active/Inactive)", "Functional Testing",
     "Staff with different statuses exist",
     "1. Select status filter (Active/Inactive)\n2. Observe results",
     "Only staff matching the selected status should be displayed"),

    ("STF_011", "Staff Pagination", "Verify pagination for large staff lists", "Functional Testing",
     "More than N staff members exist",
     "1. Navigate to Staff list\n2. Browse through pages",
     "Pagination should work correctly; each page displays N records"),

    ("STF_012", "Assign Workspace to Staff", "Verify assigning a staff member to a workspace", "Functional Testing",
     "Staff and workspaces exist",
     "1. Open staff profile\n2. Navigate to Workspace assignment\n3. Assign workspace(s)\n4. Save",
     "Staff should be assigned to selected workspace(s) successfully"),

    ("STF_013", "Staff Profile View", "Verify viewing staff profile details", "Functional Testing",
     "At least one staff member exists",
     "1. Click on a staff name\n2. Observe profile details",
     "Profile should display: name, email, phone, role, assigned workspaces, activity log"),

    ("STF_014", "Export Staff List", "Verify exporting staff data", "Functional Testing",
     "Staff list has records",
     "1. Click Export button\n2. Select format (CSV/Excel/PDF)\n3. Download",
     "Staff data should be exported in the selected format with all relevant columns"),

    ("STF_015", "Bulk Import Staff (CSV)", "Verify bulk staff import via CSV file", "Functional Testing",
     "CSV file with staff data is prepared",
     "1. Click Import\n2. Upload CSV file\n3. Map columns if needed\n4. Import",
     "Staff should be imported in bulk; success/failure report should be shown"),

    ("STF_016", "Change Staff Role", "Verify changing staff role", "Functional Testing",
     "At least one staff member with editable role exists",
     "1. Open staff edit\n2. Change role from dropdown\n3. Save",
     "Role should be updated; permissions should reflect the new role"),

    ("STF_017", "Activity Log for Staff", "Verify staff activity tracking", "Security Testing",
     "Staff member has performed actions",
     "1. Open staff profile\n2. View Activity Log",
     "All actions (login, create, edit, delete) should be logged with timestamp"),

    ("STF_018", "Password Reset for Staff", "Verify admin can reset staff password", "Functional Testing",
     "At least one active staff member exists",
     "1. Click 'Reset Password' on staff profile\n2. Confirm action",
     "Password reset link/email should be sent to staff's registered email"),

    ("STF_019", "Deactivate/Activate Staff", "Verify toggling staff active/inactive status", "Functional Testing",
     "At least one active and one inactive staff exist",
     "1. Click Deactivate on active staff\n2. Confirm\n3. Click Activate on inactive staff",
     "Staff status should toggle; deactivated staff cannot log in"),

    ("STF_020", "Permission Consistency", "Verify staff can only access permitted modules", "Security Testing",
     "Staff with limited role (e.g., Viewer) exists",
     "1. Log in as restricted staff\n2. Try to access admin-only features",
     "Restricted staff should see 403 or hidden options for unauthorized modules"),
]

for i, (tid, scenario, desc, ttype, precond, steps, expected) in enumerate(staff_cases, 2):
    row = [tid, "Staffs", scenario, desc, ttype, precond, steps, expected, "", "", ""]
    for col, val in enumerate(row, 1):
        cell = ws3.cell(row=i, column=col, value=val)
        cell.alignment = body_align
        cell.border = thin_border

for i, w in enumerate(col_widths, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w


# ============ SHEET 4: CLIENTS ============
ws4 = wb.create_sheet("Clients")

for col, h in enumerate(headers, 1):
    cell = ws4.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    cell.alignment = header_align
    cell.border = thin_border

client_cases = [
    ("CLT_001", "UI Verification", "Verify client listing page UI", "UI/UX Testing",
     "User is logged in",
     "1. Navigate to Clients module\n2. Observe the client list view",
     "Client list displays correctly with name, contact, status, and action buttons"),

    ("CLT_002", "Add Client", "Verify adding a new client", "Functional Testing",
     "User has permission to add clients",
     "1. Click 'Add Client'\n2. Fill in name, email, phone, company\n3. Click Save",
     "New client should be created and appear in the client list"),

    ("CLT_003", "Add Client - Required Fields", "Verify validation on required fields", "Negative Testing",
     "User is on Add Client form",
     "1. Leave required fields empty\n2. Click Save",
     "Validation messages: 'Client name is required', 'Email is required'"),

    ("CLT_004", "Add Client - Duplicate Email", "Verify duplicate email rejection", "Negative Testing",
     "Email already exists in system",
     "1. Enter an already registered email\n2. Fill other fields\n3. Click Save",
     "Error: 'Email already exists' should be displayed"),

    ("CLT_005", "Edit Client", "Verify editing client details", "Functional Testing",
     "At least one client exists",
     "1. Click Edit on a client\n2. Modify details\n3. Click Update",
     "Client details should be updated successfully"),

    ("CLT_006", "Delete Client", "Verify deleting a client record", "Functional Testing",
     "At least one client exists",
     "1. Click Delete on a client\n2. Confirm deletion",
     "Client should be soft-deleted / moved to trash"),

    ("CLT_007", "Search Client", "Verify client search functionality", "Functional Testing",
     "Multiple clients exist",
     "1. Enter name/email/company in search bar\n2. Observe results",
     "Clients matching the search keyword should be displayed"),

    ("CLT_008", "Filter Clients", "Verify filtering clients by status/type", "Functional Testing",
     "Clients with different statuses exist",
     "1. Select filter criteria (Active/Inactive/Lead)\n2. Observe results",
     "Only clients matching the filter criteria should be displayed"),

    ("CLT_009", "Client Profile Page", "Verify viewing full client profile", "Functional Testing",
     "At least one client exists",
     "1. Click on a client name\n2. Observe profile details",
     "Profile should show: name, email, phone, company, address, notes, project history"),

    ("CLT_010", "Assign Client to Workspace", "Verify assigning client to a workspace", "Functional Testing",
     "Client and workspaces exist",
     "1. Open client profile\n2. Go to Workspace assignment\n3. Select workspace\n4. Save",
     "Client should be linked to the selected workspace"),

    ("CLT_011", "Client Communication Log", "Verify adding/viewing communication history", "Functional Testing",
     "Client exists",
     "1. Open client profile\n2. Go to Communication Log\n3. Add a new note/call log\n4. Save",
     "Communication entry should be saved with timestamp and user details"),

    ("CLT_012", "Import Clients (CSV)", "Verify bulk import of clients", "Functional Testing",
     "CSV file with client data is prepared",
     "1. Click Import\n2. Upload CSV\n3. Map columns\n4. Import",
     "Clients should be bulk-imported; success/failure summary shown"),

    ("CLT_013", "Export Clients", "Verify exporting client list", "Functional Testing",
     "Client list has records",
     "1. Click Export\n2. Select format\n3. Download",
     "Client data exported correctly with all relevant columns"),

    ("CLT_014", "Merge Duplicate Clients", "Verify merging duplicate client records", "Functional Testing",
     "Duplicate clients exist",
     "1. Select two duplicate clients\n2. Click Merge\n3. Choose primary record\n4. Confirm",
     "Clients should be merged; data combined; duplicates archived"),

    ("CLT_015", "Client Status Change", "Verify changing client status (Lead -> Active -> Inactive)", "Functional Testing",
     "Client exists",
     "1. Open client edit\n2. Change status\n3. Save",
     "Status should be updated; any status-dependent flows should reflect the change"),

    ("CLT_016", "Add Notes/Documents to Client", "Verify attaching notes and documents", "Functional Testing",
     "Client exists",
     "1. Open client profile\n2. Go to Notes/Documents tab\n3. Add note / upload file",
     "Notes/files should be attached to client record and viewable"),

    ("CLT_017", "Client Pagination", "Verify pagination for large client lists", "Performance Testing",
     "50+ clients exist",
     "1. Navigate to client list\n2. Browse pages",
     "Pagination works; page load time is acceptable"),

    ("CLT_018", "Permission - View Only", "Verify restricted user can only view clients", "Security Testing",
     "User with Viewer role exists",
     "1. Log in as Viewer\n2. Try to Add/Edit/Delete a client",
     "Add/Edit/Delete buttons should be hidden/disabled for Viewers"),

    ("CLT_019", "Client Deletion - Active Projects", "Verify deleting a client with active projects", "Negative Testing",
     "Client has active projects assigned",
     "1. Try to delete a client with active projects",
     "Warning should appear: 'Client has active projects. Reassign or close projects first'"),

    ("CLT_020", "Data Validation - Phone/Email", "Verify phone and email format validation", "Negative Testing",
     "Client add/edit form is open",
     "1. Enter invalid phone number\n2. Enter invalid email\n3. Click Save",
     "Validation errors should appear for both fields with proper format hints"),
]

for i, (tid, scenario, desc, ttype, precond, steps, expected) in enumerate(client_cases, 2):
    row = [tid, "Clients", scenario, desc, ttype, precond, steps, expected, "", "", ""]
    for col, val in enumerate(row, 1):
        cell = ws4.cell(row=i, column=col, value=val)
        cell.alignment = body_align
        cell.border = thin_border

for i, w in enumerate(col_widths, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w


# ============ SHEET 5: SIDEBARS & NAVIGATION ============
ws5 = wb.create_sheet("Sidebars & Navigation")

sb_headers = ["Test Case ID", "Module / Sidebar", "Test Scenario", "Test Case Description", "Test Type",
              "Preconditions", "Test Steps", "Expected Result", "Actual Result", "Status (Pass/Fail)", "Remarks"]

sb_header_fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")

for col, h in enumerate(sb_headers, 1):
    cell = ws5.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = sb_header_fill
    cell.alignment = header_align
    cell.border = thin_border

sidebar_cases = [
    # ---- APP/WORKSPACE SIDEBAR (app-sidebar.tsx) ----
    ("SBR_001", "App Sidebar - General", "Verify App sidebar renders with all menu items", "UI/UX Testing",
     "User is logged in with workspace admin role",
     "1. Log in as workspace admin\n2. Observe the left sidebar",
     "Sidebar displays: Dashboard, Assign Tasks, Employees, Projects, Approvals, Time Tracker, File Manager, Billing, Interaction Followups, Inventory, Addons, Settings"),

    ("SBR_002", "App Sidebar - Dashboard", "Verify Dashboard menu item navigation", "Functional Testing",
     "User is logged in",
     "1. Click 'Dashboard' in sidebar\n2. Verify URL and page content",
     "Navigates to /dashboard; Dashboard page loads with correct content"),

    ("SBR_003", "App Sidebar - Assign Tasks", "Verify Assign Tasks menu item navigation", "Functional Testing",
     "User is logged in",
     "1. Click 'Assign Tasks' in sidebar\n2. Verify URL and page content",
     "Navigates to /overview; Assign Tasks page loads correctly"),

    ("SBR_004", "App Sidebar - Employees", "Verify Employees menu item navigation", "Functional Testing",
     "User is logged in",
     "1. Click 'Employees' in sidebar\n2. Verify URL and page content",
     "Navigates to /employees; Employee management page loads correctly"),

    ("SBR_005", "App Sidebar - Projects", "Verify Projects menu item navigation", "Functional Testing",
     "User is logged in",
     "1. Click 'Projects' in sidebar\n2. Verify URL and page content",
     "Navigates to /projects; Projects page loads correctly"),

    ("SBR_006", "App Sidebar - Approvals", "Verify Approvals menu item navigation", "Functional Testing",
     "User is logged in",
     "1. Click 'Approvals' in sidebar\n2. Verify URL and page content",
     "Navigates to /approvals; Approvals page loads correctly"),

    ("SBR_007", "App Sidebar - Time Tracker", "Verify Time Tracker menu item navigation", "Functional Testing",
     "User is logged in",
     "1. Click 'Time Tracker' in sidebar\n2. Verify URL and page content",
     "Navigates to /time-tracker; Time Tracker page loads correctly"),

    ("SBR_008", "App Sidebar - File Manager", "Verify File Manager menu item navigation", "Functional Testing",
     "User is logged in",
     "1. Click 'File Manager' in sidebar\n2. Verify URL and page content",
     "Navigates to /files; File Manager page loads correctly"),

    ("SBR_009", "App Sidebar - Billing", "Verify Billing menu item navigation", "Functional Testing",
     "User is logged in",
     "1. Click 'Billing' in sidebar\n2. Verify URL and page content",
     "Navigates to /billing; Billing page loads correctly"),

    ("SBR_010", "App Sidebar - Interaction Followups", "Verify Interaction Followups menu item navigation", "Functional Testing",
     "User is logged in",
     "1. Click 'Interaction Followups' in sidebar\n2. Verify URL and page content",
     "Navigates to /engagement; Interaction Followups page loads correctly"),

    ("SBR_011", "App Sidebar - Inventory", "Verify Inventory menu item navigation", "Functional Testing",
     "User is logged in",
     "1. Click 'Inventory' in sidebar\n2. Verify URL and page content",
     "Navigates to /stocks; Inventory page loads correctly"),

    ("SBR_012", "App Sidebar - Addons", "Verify Addons menu item navigation", "Functional Testing",
     "User is logged in",
     "1. Click 'Addons' in sidebar\n2. Verify URL and page content",
     "Navigates to /addons; Addons marketplace page loads correctly"),

    ("SBR_013", "App Sidebar - Settings", "Verify Settings menu item navigation", "Functional Testing",
     "User is logged in",
     "1. Click 'Settings' in sidebar\n2. Verify URL and page content",
     "Navigates to /settings; Settings page loads correctly"),

    ("SBR_014", "App Sidebar - Active State", "Verify active menu item highlighting", "UI/UX Testing",
     "User is logged in",
     "1. Click a sidebar menu item\n2. Observe the item's visual state",
     "Active menu item should be highlighted/indicated visually (different bg color or font weight)"),

    ("SBR_015", "App Sidebar - Collapse/Expand", "Verify sidebar collapse/expand toggle", "UI/UX Testing",
     "User is logged in",
     "1. Click the collapse/hamburger toggle\n2. Observe sidebar behavior",
     "Sidebar should collapse to icon-only view and expand back on toggle click"),

    ("SBR_016", "App Sidebar - Photography Module", "Verify Photography menu item shows conditionally", "Functional Testing",
     "Photography module is installed/not installed",
     "1. Check if Photography menu appears when module is installed\n2. Verify it hides when module is uninstalled",
     "Photography menu should only appear when the photography module API reports it as installed"),

    # ---- ORG/ORIGIN SIDEBAR (org-sidebar.tsx) ----
    ("SBR_017", "Org Sidebar - General", "Verify Org sidebar renders with all sections", "UI/UX Testing",
     "User is logged in with org_admin role",
     "1. Navigate to Org menu\n2. Observe the org sidebar",
     "Sidebar displays: Dashboard, Organization, Members, Audit Logs, Security, Plans, Blog, Settings with their sub-items"),

    ("SBR_018", "Org Sidebar - Dashboard", "Verify Org Dashboard navigation", "Functional Testing",
     "User has org_admin role",
     "1. Click 'Dashboard' in org sidebar\n2. Verify URL",
     "Navigates to /orgmenu; Org Dashboard loads correctly"),

    ("SBR_019", "Org Sidebar - Members with Sub-items", "Verify Members section expand/collapse and sub-items", "Functional Testing",
     "User has org_admin role",
     "1. Click 'Members' in org sidebar\n2. Verify sub-items appear\n3. Click 'All Members', 'Invite', 'Roles'",
     "Sub-items expand: All Members (/orgmenu/members), Invite (/orgmenu/members/invite), Roles (/orgmenu/members/roles)"),

    ("SBR_020", "Org Sidebar - Audit Logs", "Verify Audit Logs section with sub-items", "Functional Testing",
     "User has org_admin role",
     "1. Click 'Audit Logs'\n2. Verify sub-items: All Logs, Exports",
     "All Logs (/orgmenu/audit) and Exports (/orgmenu/audit/exports) navigation works"),

    ("SBR_021", "Org Sidebar - Security", "Verify Security section with sub-items", "Functional Testing",
     "User has org_admin role",
     "1. Click 'Security'\n2. Verify sub-items: Policies, SSO",
     "Policies (/orgmenu/security/policies) and SSO (/orgmenu/security/sso) navigation works"),

    ("SBR_022", "Org Sidebar - Plans", "Verify Plans navigation", "Functional Testing",
     "User has org_admin role",
     "1. Click 'Plans'\n2. Verify URL",
     "Navigates to /orgmenu/plans; Plans page loads correctly"),

    ("SBR_023", "Org Sidebar - Blog", "Verify Blog section with sub-items", "Functional Testing",
     "User has org_admin role",
     "1. Click 'Blog'\n2. Verify sub-items: All Posts, New Post",
     "All Posts (/orgmenu/blog) and New Post (/orgmenu/blog/editor) navigation works"),

    ("SBR_024", "Org Sidebar - Settings", "Verify Org Settings navigation", "Functional Testing",
     "User has org_admin role",
     "1. Click 'Settings' in org sidebar\n2. Verify URL",
     "Navigates to /orgmenu/settings; Org Settings page loads correctly"),

    # ---- STAFF SIDEBAR (staff-sidebar.tsx) ----
    ("SBR_025", "Staff Sidebar - General", "Verify Staff sidebar renders with all menu items", "UI/UX Testing",
     "User is logged in with staff role",
     "1. Log in as staff user\n2. Observe the staff sidebar",
     "Sidebar displays: Dashboard, Task, Time Sheet, Upcoming Tasks, File Management, Activity"),

    ("SBR_026", "Staff Sidebar - Dashboard", "Verify Staff Dashboard navigation", "Functional Testing",
     "User is logged in as staff",
     "1. Click 'Dashboard' in staff sidebar\n2. Verify URL",
     "Navigates to /staffs; Staff Dashboard loads correctly"),

    ("SBR_027", "Staff Sidebar - Task", "Verify Staff Task navigation", "Functional Testing",
     "User is logged in as staff",
     "1. Click 'Task' in staff sidebar\n2. Verify URL",
     "Navigates to /staffs/tasks; Staff Tasks page loads correctly"),

    ("SBR_028", "Staff Sidebar - Time Sheet", "Verify Staff Time Sheet navigation", "Functional Testing",
     "User is logged in as staff",
     "1. Click 'Time Sheet' in staff sidebar\n2. Verify URL",
     "Navigates to /staffs/timesheet; Time Sheet page loads correctly"),

    ("SBR_029", "Staff Sidebar - Upcoming Tasks", "Verify Staff Upcoming Tasks navigation", "Functional Testing",
     "User is logged in as staff",
     "1. Click 'Upcoming Tasks' in staff sidebar\n2. Verify URL",
     "Navigates to /staffs/upcoming-tasks; Upcoming Tasks page loads correctly"),

    ("SBR_030", "Staff Sidebar - File Management", "Verify Staff File Management navigation", "Functional Testing",
     "User is logged in as staff",
     "1. Click 'File Management' in staff sidebar\n2. Verify URL",
     "Navigates to /staffs/files; Staff File Manager loads correctly"),

    ("SBR_031", "Staff Sidebar - Activity", "Verify Staff Activity navigation", "Functional Testing",
     "User is logged in as staff",
     "1. Click 'Activity' in staff sidebar\n2. Verify URL",
     "Navigates to /staffs/activity; Staff Activity page loads correctly"),

    # ---- CLIENT SIDEBAR (client-sidebar.tsx) ----
    ("SBR_032", "Client Sidebar - General", "Verify Client sidebar renders with all menu items", "UI/UX Testing",
     "User is logged in with client role",
     "1. Log in as client user\n2. Observe the client sidebar",
     "Sidebar displays: Dashboard, File Management, Bills"),

    ("SBR_033", "Client Sidebar - Dashboard", "Verify Client Dashboard navigation", "Functional Testing",
     "User is logged in as client",
     "1. Click 'Dashboard' in client sidebar\n2. Verify URL",
     "Navigates to /client/dashboard; Client Dashboard loads correctly"),

    ("SBR_034", "Client Sidebar - File Management", "Verify Client File Management navigation", "Functional Testing",
     "User is logged in as client",
     "1. Click 'File Management' in client sidebar\n2. Verify URL",
     "Navigates to /client/file-manager; Client File Manager loads correctly"),

    ("SBR_035", "Client Sidebar - Bills", "Verify Client Bills navigation", "Functional Testing",
     "User is logged in as client",
     "1. Click 'Bills' in client sidebar\n2. Verify URL",
     "Navigates to /client/bills; Client Bills page loads correctly"),

    # ---- ROLE-BASED ACCESS ----
    ("SBR_036", "Role-Based Access - Admin", "Verify admin sees all app sidebar items", "Security Testing",
     "User is logged in as admin",
     "1. Observe all sidebar items\n2. Verify all items are visible",
     "Admin should see all menu items: Dashboard, Tasks, Employees, Projects, Approvals, Time Tracker, File Manager, Billing, etc."),

    ("SBR_037", "Role-Based Access - Staff", "Verify staff sees only staff sidebar (not app sidebar)", "Security Testing",
     "User is logged in as staff",
     "1. Observe the sidebar menu\n2. Verify menu items",
     "Staff should see staff-specific sidebar with Dashboard, Task, Time Sheet, Upcoming Tasks, File Management, Activity only"),

    ("SBR_038", "Role-Based Access - Client", "Verify client sees only client sidebar", "Security Testing",
     "User is logged in as client",
     "1. Observe the sidebar menu\n2. Verify menu items",
     "Client should see client-specific sidebar with Dashboard, File Management, Bills only"),

    ("SBR_039", "Role-Based Access - Org Admin", "Verify org_admin sees org sidebar", "Security Testing",
     "User is logged in as org_admin",
     "1. Observe the org sidebar menu\n2. Verify menu items with sub-menus",
     "Org admin should see full org sidebar with all sections and sub-items"),

    # ---- MOBILE BOTTOM NAVIGATION ----
    ("SBR_040", "Mobile Nav - Workspace", "Verify mobile bottom nav for workspace role", "Responsive Testing",
     "User is logged in as workspace user on mobile viewport",
     "1. Resize browser to mobile width or open on mobile device\n2. Observe bottom navigation bar",
     "Mobile bottom nav shows: Dashboard, Tasks, Time, Team, Invoices"),

    ("SBR_041", "Mobile Nav - Staff", "Verify mobile bottom nav for staff role", "Responsive Testing",
     "User is logged in as staff on mobile viewport",
     "1. Open on mobile\n2. Observe bottom nav items",
     "Staff mobile nav shows: Home, Tasks, Time, Settings"),

    ("SBR_042", "Mobile Nav - Client", "Verify mobile bottom nav for client role", "Responsive Testing",
     "User is logged in as client on mobile viewport",
     "1. Open on mobile\n2. Observe bottom nav items",
     "Client mobile nav shows: Home, Bills"),

    ("SBR_043", "Mobile Nav - Origin", "Verify mobile bottom nav for origin role", "Responsive Testing",
     "User is logged in as org user on mobile viewport",
     "1. Open on mobile\n2. Observe bottom nav items",
     "Origin mobile nav shows: Home, Members, Org, Settings"),

    ("SBR_044", "Mobile Nav - Navigation", "Verify mobile bottom nav items navigate correctly", "Functional Testing",
     "User is logged in on mobile viewport",
     "1. Tap each mobile nav item\n2. Verify page navigation",
     "Each bottom nav item navigates to the correct route and highlights as active"),

    # ---- GLOBAL UI ELEMENTS ----
    ("SBR_045", "Sidebar Responsiveness", "Verify sidebar adapts to different screen sizes", "Responsive Testing",
     "User is logged in",
     "1. Resize browser from desktop to tablet to mobile\n2. Observe sidebar behavior",
     "Desktop: full sidebar; Tablet: collapsible sidebar; Mobile: bottom navigation bar appears"),

    ("SBR_046", "User Avatar in Sidebar", "Verify user avatar/info displays in sidebar", "UI/UX Testing",
     "User is logged in",
     "1. Look at the bottom/top of sidebar\n2. Observe user info section",
     "User avatar, name, and email should be displayed in sidebar"),

    ("SBR_047", "Logout from Sidebar", "Verify logout option in sidebar", "Functional Testing",
     "User is logged in",
     "1. Click user avatar/section in sidebar\n2. Click 'Logout' option\n3. Verify behavior",
     "User should be logged out and redirected to login page"),

    ("SBR_048", "Sidebar - Keyboard Navigation", "Verify sidebar navigation via keyboard", "Usability Testing",
     "User is logged in",
     "1. Press Tab to navigate through sidebar items\n2. Press Enter to select",
     "All sidebar items should be reachable and selectable via keyboard"),
]

for i, (tid, scenario, desc, ttype, precond, steps, expected) in enumerate(sidebar_cases, 2):
    row = [tid, "Sidebars & Navigation", scenario, desc, ttype, precond, steps, expected, "", "", ""]
    for col, val in enumerate(row, 1):
        cell = ws5.cell(row=i, column=col, value=val)
        cell.alignment = body_align
        cell.border = thin_border

for i, w in enumerate(col_widths, 1):
    ws5.column_dimensions[get_column_letter(i)].width = w


# ============ SHEET 6: WORKFLOW ============
ws6 = wb.create_sheet("Workflow")

wf_header_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")

for col, h in enumerate(sb_headers, 1):
    cell = ws6.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = wf_header_fill
    cell.alignment = header_align
    cell.border = thin_border

workflow_cases = [
    # ---- APPROVALS WORKFLOW ----
    ("WFW_001", "Approvals - UI", "Verify Approvals page layout with tabs", "UI/UX Testing",
     "User is logged in with approval access",
     "1. Navigate to /approvals\n2. Observe page layout",
     "Page displays three tabs: Pending Approvals, Approved, Rejected"),

    ("WFW_002", "Approvals - Pending Tab", "Verify Pending Approvals tab shows submitted/completed items", "Functional Testing",
     "There are pending task submissions and file uploads",
     "1. Go to Approvals page\n2. Click 'Pending Approvals' tab\n3. Observe items",
     "Pending tab shows tasks with status 'submitted'/'completed' and files with status 'pending'"),

    ("WFW_003", "Approvals - Approve Task", "Verify approving a pending task", "Functional Testing",
     "At least one task is in pending approval",
     "1. Go to Pending Approvals tab\n2. Click 'Approve' on a task\n3. Add optional note\n4. Confirm",
     "Task status changes to 'approved'/'done'; task moves to Approved tab"),

    ("WFW_004", "Approvals - Reject Task", "Verify rejecting a pending task", "Functional Testing",
     "At least one task is in pending approval",
     "1. Go to Pending Approvals tab\n2. Click 'Reject' on a task\n3. Enter rejection reason\n4. Confirm",
     "Task status changes to 'cancelled'/'rejected'; task moves to Rejected tab"),

    ("WFW_005", "Approvals - Approve File", "Verify approving a pending file upload", "Functional Testing",
     "At least one file is pending approval",
     "1. Go to Pending Approvals tab\n2. Click 'Approve' on a pending file\n3. Confirm",
     "File status changes to 'approved'; file becomes accessible"),

    ("WFW_006", "Approvals - Reject File", "Verify rejecting a pending file upload", "Functional Testing",
     "At least one file is pending approval",
     "1. Go to Pending Approvals tab\n2. Click 'Reject' on a pending file\n3. Enter rejection reason\n4. Confirm",
     "File status changes to 'rejected'; file uploader is notified"),

    ("WFW_007", "Approvals - Approved Tab", "Verify Approved tab shows approved items", "Functional Testing",
     "Some items have been approved",
     "1. Click 'Approved' tab\n2. Observe listed items",
     "Approved tab shows tasks with status 'approved'/'done' and files with status 'approved'"),

    ("WFW_008", "Approvals - Rejected Tab", "Verify Rejected tab shows rejected items", "Functional Testing",
     "Some items have been rejected",
     "1. Click 'Rejected' tab\n2. Observe listed items",
     "Rejected tab shows tasks with status 'cancelled'/'rejected' and files with status 'rejected'"),

    ("WFW_009", "Approvals - Empty States", "Verify empty state messages for each tab", "UI/UX Testing",
     "No items in a particular tab",
     "1. Navigate to an empty tab (Pending/Approved/Rejected)\n2. Observe display",
     "Appropriate empty state message should be displayed (e.g., 'No pending approvals')"),

    ("WFW_010", "Approvals - Notification", "Verify user gets notified when item requires approval", "Functional Testing",
     "A task or file is submitted for approval",
     "1. Have another user submit a task for approval\n2. Check notifications as the approver",
     "Approver should receive an in-app notification about the pending approval"),

    # ---- WORKFLOW AUTOMATION ENGINE ----
    ("WFW_011", "Automation - Create Workflow", "Verify creating a new automation workflow", "Functional Testing",
     "User has admin/automation permissions",
     "1. Navigate to Workflow Automation section\n2. Click 'Create Workflow'\n3. Enter name and description\n4. Select trigger type\n5. Configure action\n6. Save",
     "Workflow should be created successfully and listed in active workflows"),

    ("WFW_012", "Automation - Task Created Trigger", "Verify workflow triggers on task.created event", "Functional Testing",
     "An active workflow with task.created trigger exists",
     "1. Create a new task\n2. Verify workflow automation executes",
     "Workflow should trigger when a task is created; action should execute"),

    ("WFW_013", "Automation - File Uploaded Trigger", "Verify workflow triggers on file.uploaded event", "Functional Testing",
     "An active workflow with file.uploaded trigger exists",
     "1. Upload a file\n2. Verify workflow automation executes",
     "Workflow should trigger when a file is uploaded; configured action should run"),

    ("WFW_014", "Automation - Schedule Trigger", "Verify schedule-based workflow trigger", "Functional Testing",
     "A scheduled workflow exists (e.g., daily report)",
     "1. Wait until scheduled time\n2. Verify workflow executes",
     "Workflow should trigger at the configured schedule (cron interval)"),

    ("WFW_015", "Automation - Notification Action", "Verify workflow sends notification action", "Functional Testing",
     "Workflow with notification action exists",
     "1. Trigger the workflow\n2. Check notifications",
     "In-app notification should be created and sent to the target user/role"),

    ("WFW_016", "Automation - Status Change Action", "Verify workflow changes task status", "Functional Testing",
     "Workflow with status_change action exists",
     "1. Trigger the workflow (e.g., task created)\n2. Verify task status after trigger",
     "Task status should change according to the workflow configuration"),

    ("WFW_017", "Automation - Assign Action", "Verify workflow reassigns tasks", "Functional Testing",
     "Workflow with assign action exists",
     "1. Trigger the workflow\n2. Verify task assignee",
     "Task should be reassigned to the configured user/role"),

    ("WFW_018", "Automation - Escalation Action", "Verify workflow escalates with priority increase", "Functional Testing",
     "Workflow with escalate action exists",
     "1. Trigger escalation condition\n2. Verify task priority and assignee",
     "Task priority should increase and be reassigned to the escalation target"),

    ("WFW_019", "Automation - Webhook Action", "Verify workflow sends webhook POST", "Integration Testing",
     "Workflow with webhook action configured to a test endpoint",
     "1. Trigger the workflow\n2. Check webhook receiver endpoint",
     "Webhook POST request should be sent to the configured URL with correct payload"),

    ("WFW_020", "Automation - Conditions", "Verify workflow conditional evaluation", "Functional Testing",
     "Workflow with conditions exists (e.g., equals, contains, greater_than)",
     "1. Trigger event that matches condition\n2. Trigger event that does NOT match condition",
     "Action executes only when condition evaluates to true; skips when false"),

    ("WFW_021", "Automation - Pause/Activate", "Verify pausing and activating a workflow", "Functional Testing",
     "An active workflow exists",
     "1. Click 'Pause' on an active workflow\n2. Verify it does not trigger\n3. Click 'Activate'\n4. Verify it triggers again",
     "Paused workflow should not execute; activated workflow should resume execution"),

    ("WFW_022", "Automation - Execution History", "Verify workflow execution logs/history", "Security Testing",
     "Workflow has been triggered multiple times",
     "1. Navigate to workflow execution history\n2. Observe log entries",
     "Each execution should be logged with trigger, action, status (success/failure), and timestamp"),

    ("WFW_023", "Automation - Edit Workflow", "Verify editing an existing workflow", "Functional Testing",
     "An existing workflow exists",
     "1. Click Edit on a workflow\n2. Modify trigger/action/conditions\n3. Save",
     "Workflow should be updated; new configuration should apply to future triggers"),

    ("WFW_024", "Automation - Delete Workflow", "Verify deleting a workflow", "Functional Testing",
     "An existing workflow exists",
     "1. Click Delete on a workflow\n2. Confirm deletion",
     "Workflow should be removed; no further executions should occur"),

    # ---- SLA ENGINE ----
    ("WFW_025", "SLA - Create SLA Definition", "Verify creating an SLA definition", "Functional Testing",
     "User has admin permissions",
     "1. Navigate to SLA settings\n2. Click 'Add SLA'\n3. Set priority level (P0-P3)\n4. Set target times for response/acknowledge/resolve/review\n5. Save",
     "SLA definition should be created and applied to matching entities"),

    ("WFW_026", "SLA - Breach Detection", "Verify SLA breach is detected and notified", "Functional Testing",
     "SLA definition exists with targets",
     "1. Create a task that exceeds SLA response time\n2. Observe behavior",
     "Breach should be detected; notification/alert should be sent to assigned user"),

    ("WFW_027", "SLA - Escalation on Breach", "Verify SLA breach triggers escalation", "Functional Testing",
     "SLA with escalation path exists",
     "1. Allow SLA to breach\n2. Verify escalation behavior",
     "Task should be escalated to the next level as per SLA configuration"),

    ("WFW_028", "SLA - Calendar Support", "Verify SLA supports 24x7 and business hours calendars", "Functional Testing",
     "SLA definitions with different calendars exist",
     "1. Create SLA with 24x7 calendar\n2. Create SLA with business hours calendar\n3. Create tasks under each",
     "24x7 SLA tracks all time; business hours SLA only tracks during configured hours"),

    # ---- FILE APPROVAL WORKFLOW (standalone) ----
    ("WFW_029", "File Approval - Submit for Approval", "Verify submitting a file for approval", "Functional Testing",
     "User has permission to upload files",
     "1. Upload a file\n2. Select 'Submit for Approval'\n3. Confirm",
     "File should be submitted with status 'pending'; approver gets notified"),

    ("WFW_030", "File Approval - Notification Flow", "Verify notification flow for file approvals", "Functional Testing",
     "A file is submitted for approval",
     "1. User A submits a file for approval\n2. Verify User B (approver) receives notification",
     "Appropriate in-app notification should be sent to the designated approver"),

    # ---- WORKFLOW STATISTICS & REASONING ----
    ("WFW_031", "Workflow Stats - Dashboard", "Verify workflow statistics display", "Functional Testing",
     "Workflows exist with execution history",
     "1. Navigate to workflow stats/reports\n2. Observe metrics",
     "Stats should display: total workflows, active workflows, total executions, success rate"),

    ("WFW_032", "Reasoning Engine - Feature Flag", "Verify Reasoning Engine appears in sidebar if applicable", "Functional Testing",
     "User is logged in",
     "1. Check sidebar for 'Reasoning Engine' menu item\n2. Verify visibility based on feature flag",
     "Reasoning Engine should appear in sidebar only when the feature is enabled/available"),

    # ---- WEBHOOK / INTEGRATION WORKFLOW ----
    ("WFW_033", "Webhook Trigger - Incoming", "Verify workflow triggers via incoming webhook", "Integration Testing",
     "A workflow with webhook trigger exists with secret token",
     "1. Send POST request to webhook URL with valid payload and token\n2. Verify workflow execution",
     "Workflow should execute upon receiving valid webhook request; invalid requests should be rejected"),

    # ---- END-TO-END WORKFLOW SCENARIOS ----
    ("WFW_034", "E2E: Task Created -> Approval -> Notification", "Verify end-to-end workflow: task creation triggers approval flow", "Integration Testing",
     "Workflow automation and approval system are configured",
     "1. User creates a task\n2. Automation detects task creation\n3. Approval request is triggered\n4. Approver approves/rejects\n5. Notification is sent",
     "Full flow executes: task created -> approval requested -> approver notified -> action taken -> final notification sent"),

    ("WFW_035", "E2E: File Upload -> Virus Scan -> Approval", "Verify file upload processing workflow", "Integration Testing",
     "File upload with virus scan and approval flow is configured",
     "1. User uploads a file\n2. Virus scan runs (detected/clean)\n3. If clean, file goes to pending approval\n4. Approver approves/rejects",
     "File goes through upload -> scan -> approval -> final status workflow correctly"),
]

for i, (tid, scenario, desc, ttype, precond, steps, expected) in enumerate(workflow_cases, 2):
    row = [tid, "Workflow", scenario, desc, ttype, precond, steps, expected, "", "", ""]
    for col, val in enumerate(row, 1):
        cell = ws6.cell(row=i, column=col, value=val)
        cell.alignment = body_align
        cell.border = thin_border

for i, w in enumerate(col_widths, 1):
    ws6.column_dimensions[get_column_letter(i)].width = w


# ============ SHEET 7: SUMMARY / TEST METRICS ============
ws5 = wb.create_sheet("Summary")

summary_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
summary_font = Font(bold=True, color="FFFFFF", size=12)
summary_align = Alignment(horizontal="center", vertical="center")

ws5.cell(row=1, column=1, value="Module").font = summary_font
ws5.cell(row=1, column=1).fill = summary_fill
ws5.cell(row=1, column=1).alignment = summary_align
ws5.cell(row=1, column=1).border = thin_border

ws5.cell(row=1, column=2, value="Total Test Cases").font = summary_font
ws5.cell(row=1, column=2).fill = summary_fill
ws5.cell(row=1, column=2).alignment = summary_align
ws5.cell(row=1, column=2).border = thin_border

ws5.cell(row=1, column=3, value="Test Types Covered").font = summary_font
ws5.cell(row=1, column=3).fill = summary_fill
ws5.cell(row=1, column=3).alignment = summary_align
ws5.cell(row=1, column=3).border = thin_border

total_all = len(login_cases)+len(workspace_cases)+len(staff_cases)+len(client_cases)+len(sidebar_cases)+len(workflow_cases)

summary_data = [
    ("Login Page", len(login_cases), "Functional, Negative, Security, UI/UX, Usability, Performance, Integration, Responsive"),
    ("Workspace", len(workspace_cases), "Functional, Negative, Security, UI/UX, Performance"),
    ("Staffs", len(staff_cases), "Functional, Negative, Security, UI/UX, Performance"),
    ("Clients", len(client_cases), "Functional, Negative, Security, UI/UX, Performance"),
    ("Sidebars & Navigation", len(sidebar_cases), "Functional, UI/UX, Security, Responsive, Usability Testing"),
    ("Workflow", len(workflow_cases), "Functional, Integration, Security, UI/UX, Negative Testing"),
    ("TOTAL", total_all, "All testing types covered"),
]

for i, (mod, total, types) in enumerate(summary_data, 2):
    ws5.cell(row=i, column=1, value=mod).border = thin_border
    ws5.cell(row=i, column=1).alignment = body_align
    ws5.cell(row=i, column=2, value=total).border = thin_border
    ws5.cell(row=i, column=2).alignment = Alignment(horizontal="center", vertical="top")
    ws5.cell(row=i, column=3, value=types).border = thin_border
    ws5.cell(row=i, column=3).alignment = body_align
    if mod == "TOTAL":
        ws5.cell(row=i, column=1).font = Font(bold=True)
        ws5.cell(row=i, column=2).font = Font(bold=True)

ws5.column_dimensions['A'].width = 16
ws5.column_dimensions['B'].width = 18
ws5.column_dimensions['C'].width = 60


filepath = "/root/myworkspace/Manual_Test_Cases.xlsx"
wb.save(filepath)
print(f"Excel file created successfully at: {filepath}")
print(f"\nSummary:")
print(f"  Login Page              : {len(login_cases)} test cases")
print(f"  Workspace               : {len(workspace_cases)} test cases")
print(f"  Staffs                  : {len(staff_cases)} test cases")
print(f"  Clients                 : {len(client_cases)} test cases")
print(f"  Sidebars & Navigation   : {len(sidebar_cases)} test cases")
print(f"  Workflow                : {len(workflow_cases)} test cases")
print(f"  TOTAL                   : {total_all} test cases")
